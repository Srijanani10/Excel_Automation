import openpyxl
from openpyxl.drawing.image import Image
from selenium import webdriver
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.microsoft import EdgeChromiumDriverManager
from PIL import Image as PILImage
import os
import time
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import json

# Path to store titles and links
data_file = "titles_and_links.json"

# Load titles and links from the JSON file
def load_titles_and_links():
    if os.path.exists(data_file):
        with open(data_file, "r") as file:
            return json.load(file)
    return {}

# Save titles and links to the JSON file
def save_titles_and_links(data):
    with open(data_file, "w") as file:
        json.dump(data, file, indent=4)

# Delete a title and its links from the JSON file
def delete_title_and_links():
    selected_title = title_combobox.get()
    if not selected_title:
        messagebox.showerror("Error", "No title selected. Please select a title.")
        return

    if selected_title in titles_and_links:
        del titles_and_links[selected_title]
        save_titles_and_links(titles_and_links)
        title_combobox['values'] = list(titles_and_links.keys())
        messagebox.showinfo("Success", "Title and links deleted successfully.")
    else:
        messagebox.showerror("Error", "Selected title not found.")

# Edit a title and its links in the JSON file
def edit_title_and_links():
    selected_title = title_combobox.get()
    if not selected_title:
        messagebox.showerror("Error", "No title selected. Please select a title.")
        return

    # Populate the entry and text widget with the selected title and links
    title_entry.delete(0, tk.END)
    title_entry.insert(0, selected_title)
    links_text.delete("1.0", tk.END)
    links_text.insert("1.0", "\n".join([link for link, _ in titles_and_links[selected_title]['links']]))

    def save_edited_title_and_links():
        new_title = title_entry.get()
        new_links = links_text.get("1.0", tk.END).strip().splitlines()
        new_cell_positions = cell_positions_text.get("1.0", tk.END).strip().splitlines()

        if not new_title or not new_links or not new_cell_positions:
            messagebox.showerror("Error", "Please provide a new title, at least one link, and corresponding cell positions.")
            return

        if len(new_links) != len(new_cell_positions):
            messagebox.showerror("Error", "The number of links and cell positions must match.")
            return

        if selected_title in titles_and_links:
            del titles_and_links[selected_title]
            titles_and_links[new_title] = {'file_path': file_path, 'links': list(zip(new_links, new_cell_positions))}
            save_titles_and_links(titles_and_links)
            title_combobox['values'] = list(titles_and_links.keys())
            messagebox.showinfo("Success", "Title and links edited successfully.")
        else:
            messagebox.showerror("Error", "Selected title not found.")

    # Change the add button to save button for editing
    add_button.config(text="Save Changes", command=save_edited_title_and_links)

# View titles and links
def view_titles_and_links():
    selected_title = title_combobox.get()
    if not selected_title:
        messagebox.showerror("Error", "No title selected. Please select a title.")
        return

    title_info = titles_and_links.get(selected_title, {})
    file_path = title_info.get('file_path', 'N/A')
    links = title_info.get('links', [])
    links_str = "\n".join([f"{link} (Cell: {cell})" for link, cell in links])
    messagebox.showinfo("View Links", f"Title: {selected_title}\nFile Path: {file_path}\nLinks:\n{links_str}")

def take_screenshot(url, output_path):
    options = webdriver.EdgeOptions()
    options.add_argument("--disable-extensions")
    options.add_argument("--user-data-dir=C:/temp/edge_profile")
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Edg/91.0.864.67"
    )

    driver = webdriver.Edge(service=EdgeService(EdgeChromiumDriverManager().install()), options=options)

    try:
        driver.get(url)
        print("Please log in manually in the browser window that has opened.")
        while True:
            try:
                WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
                break
            except Exception:
                print("Waiting for login...")

        time.sleep(5)
        driver.save_screenshot(output_path)
    except Exception as e:
        print(f"Error during screenshot capture: {e}")
    finally:
        driver.quit()

def crop_image(input_path, output_path, crop_box):
    if not os.path.exists(input_path):
        print(f"File not found: {input_path}")
        return

    with PILImage.open(input_path) as img:
        cropped_img = img.crop(crop_box)
        cropped_img.save(output_path)

# Function to create a copy of a sheet and add screenshots with custom dimensions
def add_screenshots_to_excel(file_path, sheet_name, new_sheet_name, screenshots, cell_positions):
    wb = openpyxl.load_workbook(file_path)
    original_sheet = wb[sheet_name]
    new_sheet = wb.copy_worksheet(original_sheet)
    new_sheet.title = new_sheet_name

    dimensions = {
        'B4': {'width': 96.75, 'height': 148},
        'C4': {'width': 96.75, 'height': 148},
        'B6': {'width': 96.75, 'height': 148},
        'C6': {'width': 96.75, 'height': 148}
    }

    for screenshot, cell_position in zip(screenshots, cell_positions):
        if os.path.exists(screenshot):
            col_letter = cell_position[0]
            row_number = int(cell_position[1:])

            if cell_position in dimensions:
                new_sheet.column_dimensions[col_letter].width = dimensions[cell_position]['width']
                new_sheet.row_dimensions[row_number].height = dimensions[cell_position]['height']

            img = Image(screenshot)
            img.width = new_sheet.column_dimensions[col_letter].width * 7.9804
            img.height = new_sheet.row_dimensions[row_number].height * 1.3188
            img.anchor = cell_position

            new_sheet.add_image(img)
        else:
            print(f"Screenshot file not found: {screenshot}")

    wb.save(file_path)
    
def browse_file():
    global file_path
    file_path = filedialog.askopenfilename(title="Select Excel or CSV File", filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv")])
    if file_path:
        folder_label.config(text=f"Selected File: {file_path}")
        messagebox.showinfo("File Selected", f"Detected file: {os.path.basename(file_path)}")
    else:
        messagebox.showerror("Error", "No file selected. Please select an Excel or CSV file.")

def add_title_and_links():
    global file_path
    if not file_path:
        messagebox.showerror("Error", "No file selected. Please select a file.")
        return

    title = title_entry.get()
    links = links_text.get("1.0", tk.END).strip().splitlines()
    cell_positions = cell_positions_text.get("1.0", tk.END).strip().splitlines()

    if not title or not links or not cell_positions:
        messagebox.showerror("Error", "Please provide a title, at least one link, and corresponding cell positions.")
        return

    if len(links) != len(cell_positions):
        messagebox.showerror("Error", "The number of links and cell positions must match.")
        return

    # Save the title, file path, links, and cell positions
    titles_and_links[title] = {'file_path': file_path, 'links': list(zip(links, cell_positions))}
    save_titles_and_links(titles_and_links)
    title_combobox['values'] = list(titles_and_links.keys())
    messagebox.showinfo("Success", "Title, file path, links, and cell positions added successfully.")


def create_output_folder():
    output_folder = os.path.join(os.getcwd(), "output")
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    return output_folder

# Define the run_process function here
def run_process():
    selected_title = title_combobox.get()
    if not selected_title:
        messagebox.showerror("Error", "No title selected. Please select a title.")
        return

    links_and_positions = titles_and_links[selected_title]['links']
    links = [item[0] for item in links_and_positions]
    cell_positions = [item[1] for item in links_and_positions]
    file_path = titles_and_links[selected_title]['file_path']

    wb = openpyxl.load_workbook(file_path)
    sheet_names = wb.sheetnames
    if len(sheet_names) > 1:
        sheet_name = sheet_names[1]
    else:
        sheet_name = sheet_names[0]
    new_sheet_name = f"{sheet_name} Copy"

    output_folder = create_output_folder()
    screenshot_paths = []
    for i, url in enumerate(links):
        screenshot_path = os.path.join(output_folder, f"screenshot_{i + 1}.png")
        take_screenshot(url, screenshot_path)

        crop_box = (380, 237, 1863, 900)
        cropped_screenshot_path = os.path.join(output_folder, f"cropped_screenshot_{i + 1}.png")
        crop_image(screenshot_path, cropped_screenshot_path, crop_box)
        screenshot_paths.append(cropped_screenshot_path)

    add_screenshots_to_excel(file_path, sheet_name, new_sheet_name, screenshot_paths, cell_positions)
    messagebox.showinfo("Success", "Process completed and screenshots added to the Excel file.")

root = tk.Tk()
root.title("Excel Automation Tool")
root.geometry("400x700")  # Changed the screen size to 800x600

# Create a canvas and a scrollbar
canvas = tk.Canvas(root)
scrollbar = tk.Scrollbar(root, orient="vertical", command=canvas.yview)
scrollable_frame = ttk.Frame(canvas)

scrollable_frame.bind(
    "<Configure>",
    lambda e: canvas.configure(
        scrollregion=canvas.bbox("all")
    )
)

canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
canvas.configure(yscrollcommand=scrollbar.set)

scrollbar.pack(side="right", fill="y")
canvas.pack(side="left", fill="both", expand=True)

folder_path = None

# Load existing titles and links
titles_and_links = load_titles_and_links()

title_label = tk.Label(scrollable_frame, text="Make sure the selected folder is not opened")
folder_label = tk.Label(scrollable_frame, text="No folder selected", wraplength=400)
folder_label.pack(pady=10)

browse_button = tk.Button(scrollable_frame, text="Browse File", command=browse_file)
browse_button.pack(pady=5)

title_label = tk.Label(scrollable_frame, text="Enter Title:")
title_label.pack()
title_entry = tk.Entry(scrollable_frame, width=50)
title_entry.pack(pady=5)

links_label = tk.Label(scrollable_frame, text="Enter Links (one per line):")
links_label.pack()
links_text = tk.Text(scrollable_frame, height=5, width=50)
links_text.pack(pady=5)

cell_positions_label = tk.Label(scrollable_frame, text="Enter Cell Positions (one per line):")
cell_positions_label.pack()
cell_positions_text = tk.Text(scrollable_frame, height=5, width=50)
cell_positions_text.pack(pady=5)

add_button = tk.Button(scrollable_frame, text="Add Title and Links", command=add_title_and_links)
add_button.pack(pady=5)

delete_button = tk.Button(scrollable_frame, text="Delete Title and Links", command=delete_title_and_links)
delete_button.pack(pady=5)

edit_button = tk.Button(scrollable_frame, text="Edit Title and Links", command=edit_title_and_links)
edit_button.pack(pady=5)

view_button = tk.Button(scrollable_frame, text="View Title and Links", command=view_titles_and_links)
view_button.pack(pady=5)

title_combobox_label = tk.Label(scrollable_frame, text="Select Title:")
title_combobox_label.pack()
title_combobox = ttk.Combobox(scrollable_frame, values=list(titles_and_links.keys()), state="readonly", width=47)
title_combobox.pack(pady=5)

run_button = tk.Button(scrollable_frame, text="Run", command=run_process)
run_button.pack(pady=10)

root.mainloop()